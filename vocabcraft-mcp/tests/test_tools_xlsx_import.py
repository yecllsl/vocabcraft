# tests/test_tools_xlsx_import.py
"""xlsx_import Tool 单元测试

验证 import_xlsx_vocab 真实行为:
    - 文件验证（不存在/格式错误/openpyxl缺失）
    - Excel读取（工作表不存在/缺少必需列）
    - 数据解析（空字段/多义词合并/例句解析）
    - 批量保存（成功计数/错误计数/词汇ID列表）
所有测试通过 monkeypatch 隔离数据目录到 tmp_path，不污染真实 data/。
"""

import pytest

from vocabcraft_mcp.tools.xlsx_import import import_xlsx_vocab


@pytest.fixture
def isolated_storage(tmp_path, monkeypatch):
    """隔离数据目录到 tmp_path"""
    monkeypatch.setattr("vocabcraft_mcp.tools.crud._DEFAULT_DATA_DIR", tmp_path)
    return tmp_path


def _create_xlsx(tmp_path, filename, headers, rows):
    """创建测试用 xlsx 文件"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    wb.close()
    return path


# ──────────────────────────────────────────
# 文件验证
# ──────────────────────────────────────────

def test_import_xlsx_not_exists(isolated_storage):
    """文件不存在返回错误（data/ 目录内）"""
    result = import_xlsx_vocab(str(isolated_storage / "nonexistent.xlsx"))
    assert result["success_count"] == 0
    assert result["error_count"] == 0
    assert "文件不存在" in result["errors"][0]


def test_import_xlsx_wrong_extension(isolated_storage):
    """非 .xlsx 文件返回错误"""
    path = isolated_storage / "test.txt"
    path.write_text("hello")
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert "文件格式错误" in result["errors"][0]


def test_import_xlsx_openpyxl_not_installed(isolated_storage):
    """openpyxl 未安装时抛出 ImportError"""
    import sys
    # Create a dummy .xlsx file so file validation passes
    dummy = isolated_storage / "dummy.xlsx"
    dummy.write_bytes(b"fake")

    # Temporarily remove openpyxl from sys.modules
    saved = sys.modules.pop("openpyxl", None)
    sys.modules["openpyxl"] = None  # type: ignore[assignment]
    try:
        with pytest.raises(ImportError, match="openpyxl"):
            import_xlsx_vocab(str(dummy))
    finally:
        if saved is not None:
            sys.modules["openpyxl"] = saved
        else:
            sys.modules.pop("openpyxl", None)


def test_import_xlsx_path_outside_data_rejected(isolated_storage):
    """I-2: 跨出 data/ 目录的路径被拒绝

    AGENTS.md 安全规则要求 xlsx_path resolve 后必须位于项目 data/ 目录内，
    拒绝 `..` 跨目录读取任意文件。此处构造 data 目录外的已存在文件，
    导入应返回路径越界错误而非读取。
    """
    outside = isolated_storage.parent / "outside.xlsx"
    outside.write_bytes(b"fake")
    result = import_xlsx_vocab(str(outside))
    assert result["success_count"] == 0
    assert "data" in result["errors"][0] or "目录" in result["errors"][0] or "路径" in result["errors"][0]


# ──────────────────────────────────────────
# Excel 读取
# ──────────────────────────────────────────

def test_import_xlsx_sheet_not_found(isolated_storage):
    """工作表不存在返回错误"""
    path = _create_xlsx(isolated_storage, "test.xlsx", ["word", "definitions"], [])
    result = import_xlsx_vocab(str(path), sheet_name="NonExistent")
    assert result["success_count"] == 0
    assert "工作表" in result["errors"][0]


def test_import_xlsx_missing_required_columns(isolated_storage):
    """缺少必需列返回错误"""
    path = _create_xlsx(isolated_storage, "test.xlsx", ["word", "phonetic"], [])
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert "缺少必需列" in result["errors"][0]


# ──────────────────────────────────────────
# 数据解析
# ──────────────────────────────────────────

def test_import_xlsx_empty_word(isolated_storage):
    """空词汇行被跳过"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions"],
        [
            ["", "definition1"],
            ["word2", "definition2"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert len(result["errors"]) == 1
    assert "词汇(word)为空" in result["errors"][0]


def test_import_xlsx_empty_definitions(isolated_storage):
    """空释义行被跳过"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions"],
        [
            ["word1", ""],
            ["word2", "definition2"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert len(result["errors"]) == 1
    assert "释义(definitions)为空" in result["errors"][0]


def test_import_xlsx_single_vocab(isolated_storage):
    """单个词汇成功导入"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "phonetic", "part_of_speech", "examples"],
        [
            ["hello", "你好", "/həˈloʊ/", "int.", "Hello, world!"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0
    assert len(result["imported_vocabs"]) == 1
    assert result["imported_vocabs"][0].startswith("vocab_")


def test_import_xlsx_multiple_definitions_same_word(isolated_storage):
    """同一词汇多义项合并"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "examples"],
        [
            ["bank", "银行", "I went to the bank."],
            ["bank", "河岸", "We sat by the river bank."],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0


def test_import_xlsx_examples_parsing(isolated_storage):
    """例句分号分隔正确解析"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "examples"],
        [
            ["hello", "你好", "Hello!;How are you?;你好吗"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1


def test_import_xlsx_skip_empty_rows(isolated_storage):
    """空行被跳过"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions"],
        [
            ["hello", "你好"],
            [None, None],
            ["world", "世界"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 2


def test_import_xlsx_language_override(isolated_storage):
    """行级语言覆盖"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "language"],
        [
            ["hello", "你好", "zh"],
            ["world", "世界", "en"],
        ],
    )
    result = import_xlsx_vocab(str(path), language="en")
    assert result["success_count"] == 2


def test_import_xlsx_duplicate_word_same_language(isolated_storage):
    """相同 (word, language) 行合并为一个词汇条目"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "language"],
        [
            ["hello", "你好", "en"],
            ["hello", "Hello!", "en"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0


def test_import_xlsx_duplicate_word_different_language(isolated_storage):
    """相同 word 不同 language 保持独立条目"""
    path = _create_xlsx(
        isolated_storage, "test.xlsx",
        ["word", "definitions", "language"],
        [
            ["hello", "你好", "zh"],
            ["hello", "greeting", "en"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 2
    assert result["error_count"] == 0
    assert len(result["imported_vocabs"]) == 2


def test_import_xlsx_empty_file(isolated_storage):
    """空 Excel 文件（仅表头无数据）"""
    path = _create_xlsx(isolated_storage, "test.xlsx", ["word", "definitions"], [])
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert result["error_count"] == 0
    assert result["errors"] == []
    assert result["imported_vocabs"] == []


# ──────────────────────────────────────────
# MCP 工具注册
# ──────────────────────────────────────────

def test_import_xlsx_vocab_registered_as_mcp_tool():
    """import_xlsx_vocab 已注册为 MCP 工具"""
    import asyncio

    from vocabcraft_mcp.server import mcp

    tools = asyncio.run(mcp.list_tools())
    tool_names = [tool.name for tool in tools]
    assert "import_xlsx_vocab" in tool_names


# ──────────────────────────────────────────
# 文言文实词表格式（自动检测分支）
# ──────────────────────────────────────────

def _create_classical_xlsx(tmp_path, filename, title, header, rows):
    """创建文言文实词表格式 xlsx（首行标题 + 列头 + 数据行）"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([title])
    ws.append(header)
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    wb.close()
    return path


def test_import_classical_chinese_success(isolated_storage):
    """文言文格式被自动检测并导入，例句含篇名、词性继承、例句追加均生效"""
    path = _create_classical_xlsx(
        isolated_storage,
        "classical.xlsx",
        "1. 兵 (bīng)",
        ["词性", "词义", "例句", "篇名"],
        [
            ["n.", "武器；兵器", "兵者，国之大事。", "孙子兵法"],
            ["", "士兵；军队", "可汗大点兵。", "木兰诗"],  # 词性继承上一行
            ["", "", "兵败如山倒。", "史记"],  # 无词义：例句追加到上一条释义
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0
    assert len(result["imported_vocabs"]) == 1
    vid = result["imported_vocabs"][0]
    # 校验落地数据：语言被强制为文言文
    from vocabcraft_mcp.storage import Storage

    rec = Storage(isolated_storage).load_vocab(vid)
    assert rec is not None
    assert rec.structured.language == "zh_classical"
    defs = rec.structured.definitions
    assert len(defs) == 2
    # 第二条释义聚合了第 3 行的例句
    assert any("兵败如山倒" in e for d in defs for e in d.examples)
    # 篇名拼接
    assert any("《孙子兵法》" in e for d in defs for e in d.examples)


def test_import_classical_chinese_no_data_rows(isolated_storage):
    """仅有表头无数据行时优雅返回"""
    path = _create_classical_xlsx(
        isolated_storage,
        "classical_empty.xlsx",
        "2. 错 (cuò)",
        ["词性", "词义", "例句", "篇名"],
        [],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert "未解析到任何释义" in result["errors"][0]


def test_import_xlsx_word_type_and_original_char(isolated_storage):
    """标准格式支持 word_type / original_char 列导入"""
    path = _create_xlsx(
        isolated_storage,
        "test_word_type.xlsx",
        ["word", "definitions", "language", "word_type", "original_char"],
        [
            ["之", "助词：用在主谓之间", "zh_classical", "虚词", ""],
            ["说", "喜悦", "zh_classical", "通假字", "悦"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 2
    assert result["error_count"] == 0

    from vocabcraft_mcp.storage import Storage

    storage = Storage(isolated_storage)
    for vid in result["imported_vocabs"]:
        rec = storage.load_vocab(vid)
        assert rec is not None
        if rec.structured.word == "之":
            assert rec.structured.word_type == "虚词"
        else:
            assert rec.structured.word == "说"
            assert rec.structured.word_type == "通假字"
            assert rec.structured.original_char == "悦"


def test_import_xlsx_default_word_type(isolated_storage):
    """标准格式缺省 word_type 列时默认实词"""
    path = _create_xlsx(
        isolated_storage,
        "test_default_wt.xlsx",
        ["word", "definitions", "language"],
        [["之", "助词：用在主谓之间", "zh_classical"]],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1

    from vocabcraft_mcp.storage import Storage

    rec = Storage(isolated_storage).load_vocab(result["imported_vocabs"][0])
    assert rec is not None
    assert rec.structured.word_type == "实词"
    assert rec.structured.original_char == ""


def test_import_classical_chinese_with_word_type(isolated_storage):
    """文言文格式支持"词汇类型"列导入 word_type"""
    path = _create_classical_xlsx(
        isolated_storage,
        "classical_wt.xlsx",
        "1. 之 (zhī)",
        ["词性", "词义", "例句", "篇名", "词汇类型", "本字"],
        [
            ["助词", "用于主谓之间取消句子独立性", "臣之壮也，犹不如人。", "烛之武退秦师", "虚词", ""],
            ["代词", "他、它", "使之然也。", "劝学", "", ""],  # 类型继承上一行
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0

    from vocabcraft_mcp.storage import Storage

    rec = Storage(isolated_storage).load_vocab(result["imported_vocabs"][0])
    assert rec is not None
    assert rec.structured.language == "zh_classical"
    assert rec.structured.word_type == "虚词"


def test_import_classical_chinese_with_loan_char(isolated_storage):
    """文言文格式支持"本字"列导入通假字 original_char"""
    path = _create_classical_xlsx(
        isolated_storage,
        "classical_loan.xlsx",
        "2. 说 (yuè)",
        ["词性", "词义", "例句", "篇名", "词汇类型", "本字"],
        [
            ["动词", "高兴、喜悦", "学而时习之，不亦说乎？", "论语", "通假字", "悦"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1
    assert result["error_count"] == 0

    from vocabcraft_mcp.storage import Storage

    rec = Storage(isolated_storage).load_vocab(result["imported_vocabs"][0])
    assert rec is not None
    assert rec.structured.word_type == "通假字"
    assert rec.structured.original_char == "悦"


def test_import_xlsx_invalid_word_type_reported(isolated_storage):
    """标准格式非法 word_type 被跳过并报告，不影响其他词"""
    path = _create_xlsx(
        isolated_storage,
        "test_invalid_wt.xlsx",
        ["word", "definitions", "language", "word_type"],
        [
            ["之", "助词", "zh_classical", "副词"],  # 非法类型：应为实词/虚词/通假字
            ["而", "连词", "zh_classical", "虚词"],
        ],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1  # 合法词"而"正常导入
    assert result["error_count"] == 1
    assert "词汇类型非法" in result["errors"][0]


def test_import_classical_chinese_save_failure(monkeypatch, isolated_storage):
    """保存失败时 error_count 递增并报告"""
    path = _create_classical_xlsx(
        isolated_storage,
        "classical_fail.xlsx",
        "3. 达 (dá)",
        ["词性", "词义", "例句", "篇名"],
        [["v.", "到达", "到达目的地。", "论语"]],
    )
    monkeypatch.setattr(
        "vocabcraft_mcp.tools.xlsx_import.save_vocab",
        lambda data: {"error": "mock save failed"},
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert result["error_count"] == 1
    assert "保存失败" in result["errors"][0]


# ──────────────────────────────────────────
# 边界：指定工作表 / 读取失败
# ──────────────────────────────────────────

def test_import_xlsx_named_sheet_exists(isolated_storage):
    """指定存在的工作表时正常读取"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    target = wb.create_sheet("Vocab")
    target.append(["word", "definitions"])
    target.append(["hello", "你好"])
    path = isolated_storage / "multi_sheet.xlsx"
    wb.save(path)
    wb.close()

    result = import_xlsx_vocab(str(path), sheet_name="Vocab")
    assert result["success_count"] == 1


def test_import_xlsx_read_failure(monkeypatch, isolated_storage):
    """openpyxl 读取抛异常时优雅返回错误"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["word", "definitions"])
    path = isolated_storage / "broken.xlsx"
    wb.save(path)
    wb.close()

    def _boom(*args, **kwargs):
        raise OSError("disk error")

    monkeypatch.setattr(openpyxl, "load_workbook", _boom)
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 0
    assert "读取 Excel 文件失败" in result["errors"][0]


def test_import_xlsx_standard_with_empty_header_cell(isolated_storage):
    """表头含空单元格时跳过该列而不报错"""
    path = _create_xlsx(
        isolated_storage,
        "sparse_header.xlsx",
        ["word", "", "definitions"],
        [["hello", "ignored", "你好"]],
    )
    result = import_xlsx_vocab(str(path))
    assert result["success_count"] == 1

