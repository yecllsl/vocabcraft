# src/vocabcraft_mcp/tools/xlsx_import.py
"""Excel 文件词汇导入 Tool

从 .xlsx 文件批量导入词汇到词汇学习系统。
支持两种格式:
1. 标准格式: 列名 word/phonetic/part_of_speech/definitions/examples/language
2. 文言文实词表格式: 标题行 + 词性/词义/例句/篇名 列
"""
import re
from pathlib import Path
from typing import Any

from vocabcraft_mcp.models import normalize_language, normalize_pos
from vocabcraft_mcp.tools.crud import save_vocab


def _cell_str(cell: object) -> str:
    """安全地将单元格值转为字符串，None 返回空字符串"""
    return str(cell).strip() if cell is not None else ""


_CLASSICAL_CHINESE_KEYWORDS = {"词性", "词义", "例句", "篇名"}


def _is_classical_chinese_worksheet(worksheet: Any) -> bool:
    """检测工作表是否为文言文实词表格式

    检测规则: 读取前 5 行，若某行同时包含"词性"和"词义"列名，则判定为文言文格式。
    """
    try:
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if not row:
                continue
            cells = {str(c).strip() if c else "" for c in row}
            keyword_hits = _CLASSICAL_CHINESE_KEYWORDS & cells
            if "词性" in keyword_hits and "词义" in keyword_hits:
                return True
    except Exception:  # noqa: BLE001, B110  # 检测为启发式，失败时安全回退到标准格式
        pass
    return False


def _parse_title_row(title_text: str) -> tuple[str, str]:
    """从文言文表标题行提取词汇和音标

    支持的格式:
        "61从(cóng)"       → ("从", "cóng")
        "62. 错 (cuò)"     → ("错", "cuò")
        "63. 达(dá)"       → ("达", "dá")
        "hello"            → ("hello", "")
    """
    title = str(title_text).strip() if title_text else ""
    if not title:
        return ("", "")

    # 去掉序号前缀: "61." 或 "61" 开头
    title = re.sub(r"^\d+[\.\s]*", "", title).strip()

    # 尝试提取括号内的拼音
    m = re.search(r"[（(]([^）)]+)[）)]$", title)
    phonetic = m.group(1).strip() if m else ""

    # 去掉括号部分得到词形
    word = re.sub(r"\s*[（(][^）)]*[）)]", "", title).strip()
    word = re.sub(r"\s+", "", word)  # 去掉中间空格

    return (word, phonetic)


def _find_classical_chinese_header_row(worksheet: Any) -> tuple[int, dict[str, int]]:
    """查找文言文格式的列头行，返回 (行号, 列名→列索引 映射)

    Returns:
        (row_number, column_map): 列头所在行号（1-based），列名到列索引的映射
    """
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if not row:
            continue
        column_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            cell_text = str(cell).strip() if cell else ""
            # 模糊匹配列名
            if "词性" in cell_text:
                column_map["词性"] = col_idx
            elif "词义" in cell_text:
                column_map["词义"] = col_idx
            elif "例句" in cell_text or "例" in cell_text:
                column_map["例句"] = col_idx
            elif "篇名" in cell_text:
                column_map["篇名"] = col_idx
        if "词性" in column_map and "词义" in column_map:
            return (row_idx, column_map)
    return (0, {})


def _import_classical_chinese(
    worksheet: Any,
    workbook: Any,
    xlsx_path: str,
    language: str,
) -> dict:
    """从文言文实词表格式导入词汇"""
    lang = normalize_language(language)
    # 自动覆盖为文言文语言
    lang = "zh_classical"

    # 1. 定位标题行（第一个非空行）
    title_text = ""
    for row in worksheet.iter_rows(min_row=1, max_row=3, values_only=True):
        if not row:
            continue
        first_cell = str(row[0]).strip() if row[0] else ""
        if first_cell:
            title_text = first_cell
            break

    word, phonetic = _parse_title_row(title_text)
    if not word:
        workbook.close()
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": [f"无法从文件名提取词汇，标题行: {title_text}"],
            "imported_vocabs": [],
        }

    # 2. 定位列头行
    header_row, column_map = _find_classical_chinese_header_row(worksheet)
    if header_row == 0:
        workbook.close()
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": ["未找到词性/词义列头"],
            "imported_vocabs": [],
        }

    col_pos = column_map  # {"词性": col_idx, "词义": col_idx, "例句": col_idx, "篇名": col_idx}

    # 3. 解析数据行
    definitions: list[dict] = []
    current_pos = ""  # 当前词性（行内继承）

    for _, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not row or all(cell is None for cell in row):
            continue

        # 读取各列值（使用 _cell_str 避免 str(None) → "None" 的问题）
        raw_pos = _cell_str(row[col_pos.get("词性", -1)]) if col_pos.get("词性", -1) < len(row) else ""
        raw_def = _cell_str(row[col_pos.get("词义", -1)]) if col_pos.get("词义", -1) < len(row) else ""
        raw_example = _cell_str(row[col_pos.get("例句", -1)]) if col_pos.get("例句", -1) < len(row) else ""
        raw_source = _cell_str(row[col_pos.get("篇名", -1)]) if col_pos.get("篇名", -1) < len(row) else ""

        # 跳过完全空行
        if not raw_pos and not raw_def and not raw_example and not raw_source:
            continue

        # 词性继承：空白则沿用上一行的词性
        if raw_pos:
            current_pos = normalize_pos(raw_pos)

        # 构建例句文本（含篇名）
        example_text = raw_example
        if raw_source:
            example_text = f"{raw_example} ——《{raw_source}》" if raw_example else f"《{raw_source}》"

        if raw_def:
            # 新词义行
            # 检查是否存在相同词性+词义的 definition（合并）
            existing = None
            for d in definitions:
                if d["text"] == raw_def and d.get("part_of_speech", "") == current_pos:
                    existing = d
                    break

            if existing:
                if example_text:
                    existing["examples"].append(example_text)
            else:
                definitions.append({
                    "text": raw_def,
                    "examples": [example_text] if example_text else [],
                    "part_of_speech": current_pos,
                })
        elif example_text:
            # 无词义行：例句追加到上一个 definition
            if definitions:
                definitions[-1]["examples"].append(example_text)

    workbook.close()

    if not definitions:
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": ["未解析到任何释义，请检查表格格式"],
            "imported_vocabs": [],
        }

    # 4. 聚合 top-level 词性（从所有义项中提取唯一词性，去重排序）
    seen_pos: list[str] = []
    for d in definitions:
        pos = d.get("part_of_speech", "").strip()
        if pos and pos not in seen_pos:
            seen_pos.append(pos)
    top_pos = "、".join(seen_pos)

    vocab_data = {
        "structured": {
            "word": word,
            "phonetic": phonetic,
            "part_of_speech": top_pos,
            "definitions": definitions,
            "language": lang,
        }
    }

    result = save_vocab(vocab_data)
    if "error" in result:
        # 检查是否是重复词汇错误
        if "existing_vocab_id" in result:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": [f"词汇 '{word}' 已存在 (ID: {result['existing_vocab_id']})"],
                "imported_vocabs": [],
            }
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [f"词汇 '{word}' 保存失败: {result['error']}"],
            "imported_vocabs": [],
        }

    return {
        "success_count": 1,
        "error_count": 0,
        "errors": [],
        "imported_vocabs": [result.get("vocab_id", "")],
    }


def import_xlsx_vocab(
    xlsx_path: str,
    sheet_name: str | None = None,
    language: str = "en",
) -> dict:
    """从 .xlsx 文件批量导入词汇

    支持两种格式:
    1. 标准格式: 列名 word/phonetic/part_of_speech/definitions/examples/language
    2. 文言文实词表格式: 标题行 + 词性/词义/例句/篇名 列（自动检测）

    Args:
        xlsx_path: .xlsx 文件的本地路径
        sheet_name: 工作表名称，None 则使用第一个工作表
        language: 默认语言代码（支持别名归一化，文言文格式自动覆盖为 zh_classical）

    Returns:
        包含以下字段的字典:
        - success_count: 成功导入的词汇数
        - error_count: 失败的词汇数
        - errors: 错误详情列表
        - imported_vocabs: 成功导入的词汇ID列表
    """
    lang = normalize_language(language)
    xlsx_path_obj = Path(xlsx_path)

    if not xlsx_path_obj.exists():
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": [f"文件不存在: {xlsx_path}"],
            "imported_vocabs": [],
        }

    if xlsx_path_obj.suffix.lower() != ".xlsx":
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": [f"文件格式错误: {xlsx_path}，请提供 .xlsx 文件"],
            "imported_vocabs": [],
        }

    try:
        import openpyxl  # type: ignore[import-untyped]  # noqa: WPS433 (懒加载)
    except ImportError as exc:
        raise ImportError(
            "未安装 openpyxl。请运行 `uv sync --extra xlsx` "
            "或 `uv pip install openpyxl` 后重试。"
        ) from exc

    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl raises various exceptions
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": [f"读取 Excel 文件失败: {e!s}"],
            "imported_vocabs": [],
        }

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return {
                "success_count": 0,
                "error_count": 0,
                "errors": [f"工作表 '{sheet_name}' 不存在，可用工作表: {workbook.sheetnames}"],
                "imported_vocabs": [],
            }
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    # 检测是否为文言文实词表格式
    if _is_classical_chinese_worksheet(worksheet):
        return _import_classical_chinese(worksheet, workbook, xlsx_path, language)

    headers = []
    for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        if cell:
            headers.append(str(cell).strip().lower())
        else:
            headers.append("")

    required_columns = {"word", "definitions"}
    if not required_columns.issubset(set(headers)):
        workbook.close()
        missing = required_columns - set(headers)
        return {
            "success_count": 0,
            "error_count": 0,
            "errors": [f"缺少必需列: {missing}，当前列: {headers}"],
            "imported_vocabs": [],
        }

    vocab_groups: dict[tuple[str, str], list[dict]] = {}
    errors: list[str] = []

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        row_data: dict = {}
        for header, cell in zip(headers, row, strict=False):
            if header:
                row_data[header] = cell

        word = row_data.get("word")
        def_text_raw = row_data.get("definitions")

        if not word or not str(word).strip():
            errors.append(f"行 {row_idx}: 词汇(word)为空")
            continue

        if not def_text_raw or not str(def_text_raw).strip():
            errors.append(f"行 {row_idx}: 释义(definitions)为空")
            continue

        word = str(word).strip()
        row_language = lang
        if row_data.get("language"):
            row_language = normalize_language(str(row_data["language"]))
        group_key = (word, row_language)
        if group_key not in vocab_groups:
            vocab_groups[group_key] = []
        vocab_groups[group_key].append(row_data)

    workbook.close()

    success_count = 0
    error_count = 0
    imported_vocabs: list[str] = []

    for (word, word_language), rows in vocab_groups.items():
        definitions: list[dict] = []
        phonetic = ""
        part_of_speech = ""

        for row in rows:
            if not phonetic and row.get("phonetic"):
                phonetic = str(row["phonetic"]).strip()
            if not part_of_speech and row.get("part_of_speech"):
                part_of_speech = normalize_pos(str(row["part_of_speech"]).strip())

            def_text = str(row.get("definitions", "")).strip()
            if def_text:
                examples_raw = row.get("examples", "")
                if examples_raw:
                    examples = [
                        e.strip()
                        for e in str(examples_raw).replace("；", "\n").replace(";", "\n").split("\n")
                        if e.strip()
                    ]
                else:
                    examples = []

                definitions.append({
                    "text": def_text,
                    "examples": examples,
                })

        vocab_data = {
            "structured": {
                "word": word,
                "phonetic": phonetic,
                "part_of_speech": part_of_speech,
                "definitions": definitions,
                "language": word_language,
            }
        }

        result = save_vocab(vocab_data)
        if "error" in result:
            error_count += 1
            errors.append(f"词汇 '{word}' 保存失败: {result['error']}")
        else:
            success_count += 1
            imported_vocabs.append(result.get("vocab_id", ""))

    return {
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors,
        "imported_vocabs": imported_vocabs,
    }
